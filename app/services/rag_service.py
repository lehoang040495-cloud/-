"""
RAG Service - Retrieval Augmented Generation
Uses FAISS for vector similarity search + DeepSeek for generation
"""
import os
import json
import logging
import asyncio
from functools import partial
from pathlib import Path
from typing import Optional

import numpy as np

from app.config import settings

# Force offline mode to prevent HuggingFace downloads
os.environ['HF_HUB_OFFLINE'] = '1'
os.environ['TRANSFORMERS_OFFLINE'] = '1'

logger = logging.getLogger(__name__)


class RAGService:
    def __init__(self):
        self.embedding_model = None
        self.index = None
        self.chunks = []
        self.chunk_metadata = []
        self.index_path = os.path.join(settings.VECTOR_STORE_PATH, "faiss.index")
        self.chunks_path = os.path.join(settings.VECTOR_STORE_PATH, "chunks.json")
        self._initialized = False

    def _ensure_model(self):
        if self.embedding_model is None:
            import os
            os.environ['HF_HUB_OFFLINE'] = '1'
            os.environ['TRANSFORMERS_OFFLINE'] = '1'
            from sentence_transformers import SentenceTransformer
            model_path = os.path.abspath(settings.EMBEDDING_MODEL)
            logger.info(f"Loading embedding model from: {model_path}")
            self.embedding_model = SentenceTransformer(model_path)

    def _encode_sync(self, texts: list[str]) -> np.ndarray:
        """Sync encoding - runs in thread pool"""
        self._ensure_model()
        return self.embedding_model.encode(texts, normalize_embeddings=True, show_progress_bar=False)

    async def _encode(self, texts: list[str]) -> np.ndarray:
        """Async encode - offload to thread pool"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, partial(self._encode_sync, texts))

    def _split_text(self, text: str, chunk_size: int = None, overlap: int = None) -> list[str]:
        chunk_size = chunk_size or settings.CHUNK_SIZE
        overlap = overlap or settings.CHUNK_OVERLAP
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk = text[start:end]
            if chunk.strip():
                chunks.append(chunk.strip())
            start += chunk_size - overlap
        return chunks

    def _extract_text_from_file(self, file_path: str) -> str:
        ext = Path(file_path).suffix.lower()
        text = ""

        try:
            if ext == ".txt":
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()

            elif ext == ".pdf":
                from pypdf import PdfReader
                reader = PdfReader(file_path)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

            elif ext in (".docx", ".doc"):
                from docx import Document
                doc = Document(file_path)
                for para in doc.paragraphs:
                    if para.text.strip():
                        text += para.text + "\n"

            elif ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(file_path)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " ".join(str(cell) for cell in row if cell)
                        if row_text.strip():
                            text += row_text + "\n"

            elif ext == ".json":
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                text = json.dumps(data, ensure_ascii=False, indent=2)

            elif ext == ".md":
                with open(file_path, "r", encoding="utf-8") as f:
                    text = f.read()

        except Exception as e:
            logger.error(f"Failed to extract text from {file_path}: {e}")

        return text

    async def index_document(self, file_path: str, doc_id: int, title: str = "", category: str = "通用") -> int:
        text = self._extract_text_from_file(file_path)
        if not text.strip():
            return 0

        new_chunks = self._split_text(text)
        if not new_chunks:
            return 0

        embeddings = await self._encode(new_chunks)

        if self.index is not None:
            import faiss
            self.index.add(embeddings.astype(np.float32))
        else:
            import faiss
            dimension = embeddings.shape[1]
            self.index = faiss.IndexFlatIP(dimension)
            self.index.add(embeddings.astype(np.float32))

        for i, chunk in enumerate(new_chunks):
            self.chunks.append(chunk)
            self.chunk_metadata.append({
                "doc_id": doc_id,
                "title": title,
                "category": category,
                "chunk_index": i,
            })

        self._save_index()
        self._initialized = True
        return len(new_chunks)

    async def remove_document(self, doc_id: int):
        remaining_indices = [
            i for i, meta in enumerate(self.chunk_metadata)
            if meta["doc_id"] != doc_id
        ]

        if not remaining_indices:
            self.index = None
            self.chunks = []
            self.chunk_metadata = []
            self._save_index()
            return

        remaining_chunks = [self.chunks[i] for i in remaining_indices]
        remaining_meta = [self.chunk_metadata[i] for i in remaining_indices]

        if remaining_chunks:
            embeddings = await self._encode(remaining_chunks)
            import faiss
            dimension = embeddings.shape[1]
            self.index = faiss.IndexFlatIP(dimension)
            self.index.add(embeddings.astype(np.float32))
        else:
            self.index = None

        self.chunks = remaining_chunks
        self.chunk_metadata = remaining_meta
        self._save_index()

    async def search(self, query: str, top_k: int = 5, doc_ids: Optional[list[int]] = None) -> list[dict]:
        self._load_index()

        if self.index is None or len(self.chunks) == 0:
            return []

        query_embedding = await self._encode([query])

        search_k = min(top_k * 3, len(self.chunks))
        scores, indices = self.index.search(query_embedding.astype(np.float32), search_k)

        results = []
        for score, idx in zip(scores[0], indices[0]):
            if idx < 0 or idx >= len(self.chunks):
                continue
            meta = self.chunk_metadata[idx]
            if doc_ids and meta["doc_id"] not in doc_ids:
                continue
            results.append({
                "text": self.chunks[idx],
                "score": float(score),
                "metadata": meta,
            })
            if len(results) >= top_k:
                break

        return results

    async def get_context(self, query: str, top_k: int = 5) -> str:
        results = await self.search(query, top_k)
        if not results:
            return ""
        context_parts = [r["text"] for r in results]
        return "\n\n---\n\n".join(context_parts)

    def _save_index(self):
        import faiss
        if self.index is not None:
            faiss.write_index(self.index, self.index_path)
        data = {
            "chunks": self.chunks,
            "metadata": self.chunk_metadata,
        }
        with open(self.chunks_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _load_index(self):
        if self._initialized:
            return

        import faiss
        if os.path.exists(self.index_path) and os.path.exists(self.chunks_path):
            self.index = faiss.read_index(self.index_path)
            with open(self.chunks_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.chunks = data.get("chunks", [])
            self.chunk_metadata = data.get("metadata", [])
            self._initialized = True
            logger.info(f"Loaded {len(self.chunks)} chunks from disk")

    async def reindex_all(self):
        self.index = None
        self.chunks = []
        self.chunk_metadata = []
        self._initialized = False
        self._load_index()


rag_service = RAGService()
